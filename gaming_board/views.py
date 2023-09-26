from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
import openpyxl
import pyrebase

points = {
    "Hackathon": [100, 200, 300],
    "Shark Tank": [100, 200, 300],
    "Product Booth": [100, 200, 300],
    "Special Round": [1000, 800, 600, 400]
}

total_members = {
    "App Volumes": 31,
    "CART": 25,
    "Horizon Cloud Services": 95,
    "Horizon Enterprise and Customer Engineering": 129,
    "PM, OCTO and Mobile Ops": 34,
    "SDK, Productivity Apps and Hub Services": 88,
    "Unified Endpoint Management": 181,
    "WS1 Access": 76,
    "WS1 Assist": 16,
    "WS1 Intelligence and GEM": 54
}

last_updated_data = "Round 2, 20th September 2023, 01:58 AM"

last_updated_data_v2 = "Round 2, 24th September 2023, 10:25 PM"

firebaseConfig = {
    'apiKey': "AIzaSyD3V7p2Y897rZhIYVVUpBkUXQDpkqsJD_k",
    'authDomain': "euc-innovation-dashboard.firebaseapp.com",
    'databaseURL': "https://euc-innovation-dashboard-default-rtdb.firebaseio.com",
    'projectId': "euc-innovation-dashboard",
    'storageBucket': "euc-innovation-dashboard.appspot.com",
    'messagingSenderId': "74853721598",
    'appId': "1:74853721598:web:59da95461d07fcb9330caf"
}


def noquote(s):
    return s


firebase = pyrebase.initialize_app(firebaseConfig)
db = firebase.database()
pyrebase.pyrebase.quote = noquote


def get_points(event_name, event_round):
    return points[event_name][event_round - 1]


def get_special_round_points(array_list):
    total_point = 0
    for (index, value) in enumerate(array_list):
        total_point = total_point + (points["Special Round"][index] * value)
    return total_point


def add_update_team_score(team_name, team_score, event_name, event_round, total_participants):
    try:
        get_team_details = db.child("innovation_week").order_by_child("primary_key").equal_to(
            team_name + event_name).get()
        if len(get_team_details.each()) == 0:

            data = {
                'primary_key': team_name + event_name,
                'day_wise_event_score': [team_score],
                'total_score': team_score,
                'total_participants': total_participants,
                'team_name': team_name
            }
            db.child("innovation_week").push(data)

        else:
            team_details = get_team_details.val()
            for key, value in team_details.items():

                team_details_score = value['day_wise_event_score']
                if len(team_details_score) >= event_round:
                    previous_total_score = team_details_score[event_round - 1]
                    team_details_score[event_round - 1] = team_score

                    total_score = value['total_score'] + team_score - previous_total_score
                    participants = (team_score - previous_total_score) / get_points(event_name, event_round)
                    total_participants = int(value['total_participants'] + participants)
                    if total_score < 0:
                        total_score = 0
                    data = {
                        'primary_key': team_name + event_name,
                        'day_wise_event_score': team_details_score,
                        'total_score': total_score,
                        'total_participants': total_participants,
                        'team_name': team_name
                    }
                    db.child("innovation_week").child(str(key)).update(data)
                else:
                    total_score = value['total_score'] + team_score
                    day_wise_event_score = value['day_wise_event_score']
                    day_wise_event_score.append(team_score)
                    total_participants = value['total_participants'] + total_participants
                    data = {
                        'primary_key': team_name + event_name,
                        'day_wise_event_score': day_wise_event_score,
                        'total_score': total_score,
                        'total_participants': total_participants,
                        'team_name': team_name
                    }
                    db.child("innovation_week").child(str(key)).update(data)
        return "data added"

    except Exception as e:
        print(e)
        return "data not added due to exception:- "


def add_update_team_score_v2(team_name, team_score, event_name, event_round, total_participants):
    try:
        get_team_details = db.child("innovation_week_v2").order_by_child("primary_key").equal_to(
            team_name + event_name).get()
        if len(get_team_details.each()) == 0:

            data = {
                'primary_key': team_name + event_name,
                'day_wise_event_score': [team_score],
                'total_score': team_score,
                'total_participants': total_participants,
                'team_name': team_name
            }
            db.child("innovation_week_v2").push(data)

        else:
            team_details = get_team_details.val()
            for key, value in team_details.items():

                team_details_score = value['day_wise_event_score']
                if len(team_details_score) >= event_round:
                    previous_total_score = team_details_score[event_round - 1]
                    team_details_score[event_round - 1] = team_score

                    total_score = value['total_score'] + team_score - previous_total_score
                    participants = (team_score - previous_total_score) / get_points(event_name, event_round)
                    total_participants = int(value['total_participants'] + participants)
                    if total_score < 0:
                        total_score = 0
                    data = {
                        'primary_key': team_name + event_name,
                        'day_wise_event_score': team_details_score,
                        'total_score': total_score,
                        'total_participants': total_participants,
                        'team_name': team_name
                    }
                    db.child("innovation_week_v2").child(str(key)).update(data)
                else:
                    total_score = value['total_score'] + team_score
                    day_wise_event_score = value['day_wise_event_score']
                    day_wise_event_score.append(team_score)
                    total_participants = value['total_participants'] + total_participants
                    data = {
                        'primary_key': team_name + event_name,
                        'day_wise_event_score': day_wise_event_score,
                        'total_score': total_score,
                        'total_participants': total_participants,
                        'team_name': team_name
                    }
                    db.child("innovation_week_v2").child(str(key)).update(data)
        return "data added"

    except Exception as e:
        print(e)
        return "data not added due to exception:- "


def save_data(worksheet, event_round):
    first_row = True
    ordering = []
    for row in worksheet.iter_rows():
        first_column = True
        temp_count = 0
        team_name = ""
        for row_value in row:
            if first_row:
                ordering.append(str(row_value.value))
            elif first_column:
                first_column = False
                team_name = str(row_value.value)
            else:
                temp_count = temp_count + 1
                event_name = ordering[temp_count]
                if row_value.value is None:
                    continue
                team_score = int(row_value.value) * get_points(event_name, event_round)
                if add_update_team_score(team_name, team_score, event_name, event_round,
                                         row_value.value) != "data added":
                    return "data not added"
        first_row = False
    return "data added"


def save_data_v2(worksheet, event_round):
    first_row = True
    ordering = []
    for row in worksheet.iter_rows():
        first_column = True
        temp_count = 0
        team_name = ""
        for row_value in row:
            if first_row:
                ordering.append(str(row_value.value))
            elif first_column:
                first_column = False
                team_name = str(row_value.value)
            else:
                temp_count = temp_count + 1
                event_name = ordering[temp_count]
                if row_value.value is None:
                    continue
                team_score = int(row_value.value) * get_points(event_name, event_round)
                if add_update_team_score_v2(team_name, team_score, event_name, event_round,
                                            row_value.value) != "data added":
                    return "data not added"
        first_row = False
    return "data added"


def special_round(worksheet, event_name):
    first_row = True
    for row in worksheet.iter_rows():
        if first_row:
            first_row = False
            continue
        first_column = True
        count = []
        team_name = ""
        total_participants = 0
        for row_value in row:
            if first_column:
                first_column = False
                team_name = str(row_value.value)
            else:
                total_participants = total_participants + int(row_value.value)
                count.append(int(row_value.value))
        total_score = get_special_round_points(count)
        if add_update_team_score(team_name, total_score, event_name, 4, total_participants) != "data added":
            return "data not added"

    return "data added"


def special_round_v2(worksheet, event_name):
    first_row = True
    for row in worksheet.iter_rows():
        if first_row:
            first_row = False
            continue
        first_column = True
        count = []
        team_name = ""
        total_participants = 0
        for row_value in row:
            if first_column:
                first_column = False
                team_name = str(row_value.value)
            else:
                total_participants = total_participants + int(row_value.value)
                count.append(int(row_value.value))
        total_score = get_special_round_points(count)
        if add_update_team_score_v2(team_name, total_score, event_name, 4, total_participants) != "data added":
            return "data not added"

    return "data added"


def get_worksheet_name(event_round):
    if event_round == 1:
        return "Round_1"
    elif event_round == 2:
        return "Round_2"
    elif event_round == 3:
        return "Round_3"
    elif event_round == 4:
        return "Final_1"
    elif event_round == 5:
        return "Final_2"
    else:
        return "Final_3"


class ReadExcelFile(APIView):

    def post(self, request, pk):
        excel_file = request.FILES["EUC_Innovation_Week_2023"]
        work_book = openpyxl.load_workbook(excel_file)
        event_round = pk
        worksheet = work_book[get_worksheet_name(event_round)]
        if pk == 4:
            if special_round(worksheet, "Hackathon") == "data not added":
                return Response("Data not added", status=status.HTTP_400_BAD_REQUEST)
        elif pk == 5:
            if special_round(worksheet, "Shark Tank") == "data not added":
                return Response("Data not added", status=status.HTTP_400_BAD_REQUEST)
        elif pk == 6:
            if special_round(worksheet, "Product Booth") == "data not added":
                return Response("Data not added", status=status.HTTP_400_BAD_REQUEST)
        elif save_data(worksheet, event_round) == "data not added":
            return Response("Data not added", status=status.HTTP_400_BAD_REQUEST)

        return Response("Data Added Successfully", status=status.HTTP_201_CREATED)


class ReadExcelFileV2(APIView):

    def post(self, request, pk):
        excel_file = request.FILES["EUC_Innovation_Week_2023"]
        work_book = openpyxl.load_workbook(excel_file)
        event_round = pk
        worksheet = work_book[get_worksheet_name(event_round)]
        if pk == 4:
            if special_round_v2(worksheet, "Hackathon") == "data not added":
                return Response("Data not added", status=status.HTTP_400_BAD_REQUEST)
        elif pk == 5:
            if special_round_v2(worksheet, "Shark Tank") == "data not added":
                return Response("Data not added", status=status.HTTP_400_BAD_REQUEST)
        elif pk == 6:
            if special_round_v2(worksheet, "Product Booth") == "data not added":
                return Response("Data not added", status=status.HTTP_400_BAD_REQUEST)
        elif save_data_v2(worksheet, event_round) == "data not added":
            return Response("Data not added", status=status.HTTP_400_BAD_REQUEST)

        return Response("Data Added Successfully", status=status.HTTP_201_CREATED)


class GetTeamScore(APIView):

    def get(self, request):
        result = db.child("innovation_week").get().val()
        final_result = {

        }
        temp_dict = {}

        if result is None:
            return Response(final_result, status=status.HTTP_200_OK)

        for (key, value) in result.items():
            if value['team_name'] in temp_dict:
                temp_dict1 = temp_dict[value['team_name']]
                temp_dict1['total_score'] = temp_dict1['total_score'] + value['total_score']
                temp_dict1['total_participants'] = temp_dict1['total_participants'] + value['total_participants']

            else:
                temp_dict1 = {'product_name': value['team_name'], 'total_score': value['total_score'],
                              'total_participants': value['total_participants'],
                              'head_count': total_members[value['team_name']]}
                temp_dict[value['team_name']] = temp_dict1

        array_list = []

        for (key, value) in temp_dict.items():
            value['normalized_score'] = round((value['total_score'] * 100) / total_members[value['product_name']])
            array_list.append(value)

        array_list = sorted(array_list, key=lambda x: x["normalized_score"], reverse=True)
        rank = 1
        for data in array_list:
            data['rank'] = rank
            rank = rank + 1
        final_result['contents'] = array_list
        final_result['last_updated'] = last_updated_data
        return Response(final_result, status=status.HTTP_200_OK)


class GetTeamScoreV2(APIView):

    def get(self, request):
        result = db.child("innovation_week_v2").get().val()
        final_result = {

        }
        temp_dict = {}

        if result is None:
            return Response(final_result, status=status.HTTP_200_OK)

        for (key, value) in result.items():
            if value['team_name'] in temp_dict:
                temp_dict1 = temp_dict[value['team_name']]
                temp_dict1['total_score'] = temp_dict1['total_score'] + value['total_score']
                temp_dict1['total_participants'] = temp_dict1['total_participants'] + value['total_participants']

            else:
                temp_dict1 = {'product_name': value['team_name'], 'total_score': value['total_score'],
                              'total_participants': value['total_participants'],
                              'head_count': total_members[value['team_name']]}
                temp_dict[value['team_name']] = temp_dict1

        array_list = []

        for (key, value) in temp_dict.items():
            value['normalized_score'] = round((value['total_score'] * 100) / total_members[value['product_name']])
            array_list.append(value)

        array_list = sorted(array_list, key=lambda x: x["normalized_score"], reverse=True)
        rank = 1
        for data in array_list:
            data['rank'] = rank
            rank = rank + 1
        final_result['contents'] = array_list
        final_result['last_updated'] = last_updated_data_v2
        return Response(final_result, status=status.HTTP_200_OK)
